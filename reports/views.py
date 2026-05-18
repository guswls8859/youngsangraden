import datetime
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.utils import timezone
from django.contrib.auth import get_user_model
from .models import DailyReport, TaskItem, DailyTask, SubTask, OperationsDailyData
from .forms import DailyReportForm, DailyTaskForm
from .pdf import build_report_pdf, build_daily_pdf, build_weekly_pdf, build_daily_task_pdf, build_weekly_task_pdf


class OperationsAccessMixin(LoginRequiredMixin):
    """운영사무국 소속만 접근 가능"""
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.organization != 'operations':
            return redirect('main_menu')
        return super().dispatch(request, *args, **kwargs)


class ManagerRequiredMixin(OperationsAccessMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.role == 'manager'


def _require_operations(request):
    """운영사무국 소속이 아니면 True 반환 (리다이렉트 필요)."""
    return request.user.organization != 'operations'


def save_task_items(request, report):
    """POST 데이터에서 업무 항목을 파싱해 저장한다."""
    report.task_items.all().delete()
    count = int(request.POST.get('task_count', 0))
    for i in range(count):
        content = request.POST.get(f'task_{i}_content', '').strip()
        category = request.POST.get(f'task_{i}_category', '')
        try:
            progress = int(request.POST.get(f'task_{i}_progress', 0))
        except ValueError:
            progress = 0
        if content and category in ('completed', 'in_progress', 'tomorrow'):
            TaskItem.objects.create(
                report=report,
                content=content,
                category=category,
                progress=max(0, min(100, progress)),
                order=i,
            )


class ReportListView(OperationsAccessMixin, ListView):
    model = DailyReport
    template_name = 'reports/list.html'
    context_object_name = 'reports'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        if user.role == 'manager':
            return DailyReport.objects.select_related('author').all()
        return DailyReport.objects.filter(author=user)


class ReportDetailView(OperationsAccessMixin, DetailView):
    model = DailyReport
    template_name = 'reports/detail.html'
    context_object_name = 'report'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        report = self.object
        ctx['completed_tasks'] = report.task_items.filter(category='completed')
        ctx['in_progress_tasks'] = report.task_items.filter(category='in_progress')
        ctx['tomorrow_tasks'] = report.task_items.filter(category='tomorrow')
        return ctx


class ReportCreateView(OperationsAccessMixin, CreateView):
    model = DailyReport
    form_class = DailyReportForm
    template_name = 'reports/form.html'
    success_url = reverse_lazy('reports:list')

    def get_initial(self):
        return {'report_date': timezone.localdate()}

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.instance.status = 'submitted'
        report = form.save()
        save_task_items(self.request, report)
        return redirect(self.success_url)


class ReportUpdateView(OperationsAccessMixin, UpdateView):
    model = DailyReport
    form_class = DailyReportForm
    template_name = 'reports/form.html'
    success_url = reverse_lazy('reports:list')

    def get_queryset(self):
        return DailyReport.objects.filter(author=self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['existing_tasks'] = list(
            self.object.task_items.values('category', 'content', 'progress', 'order')
        )
        return ctx

    def form_valid(self, form):
        report = form.save()
        save_task_items(self.request, report)
        return redirect(self.success_url)


@login_required
def report_draft_save(request, pk):
    if _require_operations(request):
        return redirect('main_menu')
    report = DailyReport.objects.get(pk=pk, author=request.user)
    report.status = 'draft'
    report.save()
    return redirect('reports:detail', pk=pk)


@login_required
def report_pdf_download(request, pk):
    if _require_operations(request):
        return redirect('main_menu')
    report = get_object_or_404(
        DailyReport.objects.prefetch_related('task_items').select_related('author'),
        pk=pk
    )
    # 본인 또는 관리자만 다운로드 가능
    if report.author != request.user and request.user.role != 'manager':
        return redirect('reports:list')

    pdf_bytes = build_report_pdf(report)
    filename = f'report_{report.report_date}_{report.author.username}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def daily_pdf_download(request, date_str):
    if _require_operations(request) or request.user.role != 'manager':
        return redirect('main_menu')
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        return redirect('reports:dashboard_daily')

    reports = (
        DailyReport.objects
        .filter(report_date=target_date)
        .select_related('author')
        .prefetch_related('task_items')
        .order_by('author__last_name')
    )
    pdf_bytes = build_daily_pdf(reports, target_date)
    filename = f'daily_report_{target_date}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def weekly_pdf_download(request, week_str):
    if _require_operations(request) or request.user.role != 'manager':
        return redirect('main_menu')
    try:
        year, week_num = week_str.split('-W')
        week_start = datetime.datetime.strptime(f'{year}-W{int(week_num)}-1', '%Y-W%W-%w').date()
    except (ValueError, AttributeError):
        return redirect('reports:dashboard_weekly')

    week_end = week_start + datetime.timedelta(days=6)
    reports = (
        DailyReport.objects
        .filter(report_date__range=(week_start, week_end))
        .select_related('author')
        .prefetch_related('task_items')
        .order_by('report_date', 'author__last_name')
    )

    days = {}
    for d in range(7):
        days[week_start + datetime.timedelta(days=d)] = []
    for report in reports:
        if report.report_date in days:
            days[report.report_date].append(report)

    pdf_bytes = build_weekly_pdf(days, week_start, week_end)
    filename = f'weekly_report_{week_start}_{week_end}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── DailyTask 뷰 ─────────────────────────────────────────────

class TaskListView(OperationsAccessMixin, ListView):
    """개인 투두리스트 (본인 업무만)"""
    model = DailyTask
    template_name = 'reports/task_list.html'
    context_object_name = 'tasks'

    def get_queryset(self):
        return DailyTask.objects.filter(user=self.request.user).order_by('status', '-start_date')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['form'] = DailyTaskForm(initial={'start_date': timezone.localdate()})
        ctx['today'] = timezone.localdate()
        tasks = self.get_queryset()
        ctx['has_pending'] = tasks.exclude(status='done').exists()
        ctx['has_done'] = tasks.filter(status='done').exists()
        return ctx


@login_required
def task_create(request):
    if _require_operations(request):
        return redirect('main_menu')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if request.method == 'POST':
        form = DailyTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            end_date_str = request.POST.get('end_date', '').strip()
            if end_date_str:
                try:
                    task.end_date = datetime.date.fromisoformat(end_date_str)
                except ValueError:
                    pass
            task.save()
            # 서브업무 동시 생성
            subtask_titles = request.POST.getlist('subtask_titles')
            for i, title in enumerate(t.strip() for t in subtask_titles):
                if title:
                    SubTask.objects.create(daily_task=task, title=title, order=i)
            if any(t.strip() for t in subtask_titles):
                task.recalculate_progress()
                task.refresh_from_db()
            if is_ajax:
                return JsonResponse({
                    'ok': True, 'pk': task.pk,
                    'task_name': task.task_name,
                    'progress': task.progress,
                    'status': task.status,
                    'status_display': task.get_status_display(),
                    'note': task.note,
                })
        elif is_ajax:
            return JsonResponse({'ok': False}, status=400)
    return redirect('reports:task_calendar')


@login_required
def task_update_progress(request, pk):
    """AJAX: 진행도 슬라이더 실시간 업데이트"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    task = get_object_or_404(DailyTask, pk=pk, user=request.user)
    if request.method == 'POST':
        try:
            progress = int(request.POST.get('progress', task.progress))
        except ValueError:
            progress = task.progress
        task.progress = progress
        task.save()  # save()에서 status 자동 처리
        return JsonResponse({
            'status': task.status,
            'status_display': task.get_status_display(),
            'progress': task.progress,
        })
    return JsonResponse({'error': 'invalid'}, status=400)


@login_required
def task_update_status(request, pk):
    """상태 변경 — AJAX/일반 모두 지원"""
    if _require_operations(request):
        return redirect('main_menu')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    task = get_object_or_404(DailyTask, pk=pk, user=request.user)
    if request.method == 'POST':
        status = request.POST.get('status')
        if status in ('doing', 'hold', 'done'):
            if status == 'done':
                task.progress = 100
                task.status = 'done'
                task.save()  # save()가 completed_date 자동 처리
            else:
                # 완료→진행중/보류 전환: save()의 progress==100 강제 잠금을 우회
                update_fields = {'status': status, 'completed_date': None}
                # 서브 업무 없이 100%인 경우 진행률도 초기화
                if task.progress == 100 and not task.subtasks.exists():
                    update_fields['progress'] = 0
                DailyTask.objects.filter(pk=task.pk).update(**update_fields)
                task.refresh_from_db()
        if is_ajax:
            return JsonResponse({
                'ok': True, 'status': task.status,
                'status_display': task.get_status_display(),
                'progress': task.progress,
            })
    return redirect('reports:task_calendar')


@login_required
def task_edit(request, pk):
    """AJAX: 업무 내용 수정 (업무명, 목표완료일, 비고)"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    task = get_object_or_404(DailyTask, pk=pk, user=request.user)
    if request.method == 'POST':
        task_name = request.POST.get('task_name', '').strip()
        note = request.POST.get('note', '').strip()
        end_date_str = request.POST.get('end_date', '').strip()
        if not task_name:
            return JsonResponse({'ok': False, 'error': '업무명 필요'}, status=400)
        task.task_name = task_name
        task.note = note
        if end_date_str:
            try:
                task.end_date = datetime.date.fromisoformat(end_date_str)
            except ValueError:
                pass
        else:
            task.end_date = None
        task.save()
        return JsonResponse({
            'ok': True,
            'task_name': task.task_name,
            'note': task.note,
            'end_date': task.end_date.isoformat() if task.end_date else '',
        })
    return JsonResponse({'error': 'method'}, status=405)


@login_required
def task_delete(request, pk):
    if _require_operations(request):
        return redirect('main_menu')
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    task = get_object_or_404(DailyTask, pk=pk, user=request.user)
    if request.method == 'POST':
        task.delete()
        if is_ajax:
            return JsonResponse({'ok': True})
    return redirect('reports:task_calendar')


@login_required
def task_review_toggle(request, pk):
    """관리자 전용: 100% 완료 업무 검토 완료 토글 (AJAX POST)"""
    if request.user.organization != 'operations' or request.user.role != 'manager':
        return JsonResponse({'error': 'forbidden'}, status=403)
    task = get_object_or_404(DailyTask, pk=pk, status='done')
    if request.method == 'POST':
        if task.is_reviewed:
            task.is_reviewed = False
            task.reviewed_by = None
            task.reviewed_at = None
        else:
            task.is_reviewed = True
            task.reviewed_by = request.user
            task.reviewed_at = timezone.now()
        task.save(update_fields=['is_reviewed', 'reviewed_by', 'reviewed_at'])
        return JsonResponse({
            'ok': True,
            'is_reviewed': task.is_reviewed,
            'reviewed_by': task.reviewed_by.get_full_name() if task.reviewed_by else '',
            'reviewed_at': task.reviewed_at.strftime('%m/%d %H:%M') if task.reviewed_at else '',
        })
    return JsonResponse({'error': 'method'}, status=405)


# ── SubTask 뷰 ─────────────────────────────────────────────────

@login_required
def subtask_edit(request, pk):
    """서브 업무 제목 수정"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    subtask = get_object_or_404(SubTask, pk=pk, daily_task__user=request.user)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        if not title:
            return JsonResponse({'ok': False}, status=400)
        subtask.title = title
        subtask.save()
        return JsonResponse({'ok': True, 'title': subtask.title})
    return JsonResponse({'error': 'method'}, status=405)


@login_required
def subtask_create(request, pk):
    """서브 업무 생성"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    task = get_object_or_404(DailyTask, pk=pk, user=request.user)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        if not title:
            return JsonResponse({'ok': False, 'error': '제목 필요'}, status=400)
        order = task.subtasks.count()
        SubTask.objects.create(daily_task=task, title=title, order=order)
        task.recalculate_progress()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'method'}, status=405)


@login_required
def subtask_toggle(request, pk):
    """서브 업무 완료 토글"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    subtask = get_object_or_404(SubTask, pk=pk, daily_task__user=request.user)
    if request.method == 'POST':
        subtask.is_done = not subtask.is_done
        subtask.save()
        subtask.daily_task.recalculate_progress()
        task = DailyTask.objects.get(pk=subtask.daily_task_id)
        return JsonResponse({
            'ok': True,
            'is_done': subtask.is_done,
            'progress': task.progress,
            'status': task.status,
            'status_display': task.get_status_display(),
        })
    return JsonResponse({'error': 'method'}, status=405)


@login_required
def subtask_delete(request, pk):
    """서브 업무 삭제"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    subtask = get_object_or_404(SubTask, pk=pk, daily_task__user=request.user)
    if request.method == 'POST':
        task = subtask.daily_task
        subtask.delete()
        task.recalculate_progress()
        return JsonResponse({'ok': True})
    return JsonResponse({'error': 'method'}, status=405)


import calendar as _calendar

class TaskCalendarView(OperationsAccessMixin, TemplateView):
    """운영사무국 팀/개인 캘린더 — 날짜 클릭 시 투두 패널"""
    template_name = 'reports/task_calendar.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.localdate()
        try:
            year  = int(self.request.GET.get('year',  today.year))
            month = int(self.request.GET.get('month', today.month))
            if not (1 <= month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            year, month = today.year, today.month

        mode = self.request.GET.get('mode', 'team')
        if mode not in ('team', 'personal'):
            mode = 'team'

        prev_year,  prev_month  = (year - 1, 12) if month == 1  else (year, month - 1)
        next_year,  next_month  = (year + 1, 1)  if month == 12 else (year, month + 1)

        cal = _calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)

        import calendar as _cal_mod
        last_day = _cal_mod.monthrange(year, month)[1]
        month_start = datetime.date(year, month, 1)
        month_end = datetime.date(year, month, last_day)

        if mode == 'personal':
            # 이달 범위와 겹치는 내 업무 전부
            tasks_qs = (
                DailyTask.objects
                .filter(
                    user=self.request.user,
                    user__organization='operations',
                    start_date__lte=month_end,
                )
                .filter(
                    Q(end_date__gte=month_start) |
                    Q(end_date__isnull=True, start_date__gte=month_start)
                )
                .select_related('user')
            )
        else:
            tasks_qs = (
                DailyTask.objects
                .filter(start_date__year=year, start_date__month=month,
                        user__organization='operations')
                .select_related('user')
            )

        import json
        day_map: dict = {}

        if mode == 'personal':
            # 개인: 업무명을 날짜별로 나열
            for t in tasks_qs:
                task_info = {'task_name': t.task_name, 'status': t.status}
                if t.end_date:
                    span_start = max(t.start_date, month_start)
                    span_end   = min(t.end_date, month_end)
                    cur = span_start
                    while cur <= span_end:
                        day_map.setdefault(cur.day, []).append(task_info)
                        cur += datetime.timedelta(days=1)
                elif month_start <= t.start_date <= month_end:
                    day_map.setdefault(t.start_date.day, []).append(task_info)
            day_dots_json = json.dumps(day_map)
        else:
            # 팀: 사용자별 집계
            team_map: dict = {}
            for t in tasks_qs:
                uid = t.user_id
                d = t.start_date.day
                team_map.setdefault(d, {})
                if uid not in team_map[d]:
                    team_map[d][uid] = {
                        'name': t.user.get_full_name() or t.user.username,
                        'emoji': t.user.emoji,
                        'total': 0, 'done': 0,
                    }
                team_map[d][uid]['total'] += 1
                if t.status == 'done':
                    team_map[d][uid]['done'] += 1
            day_dots_json = json.dumps({
                d: list(v.values()) for d, v in team_map.items()
            })

        emoji_list = [
            '☃️',
            # 동물
            '🐶','🐱','🐭','🐰','🦊','🐻','🐼','🐨','🐯','🦁',
            '🐸','🐵','🐔','🐧','🐦','🦆','🦅','🦉','🦋','🐝',
            # 자연·식물
            '🌿','🌱','🌲','🌳','🌴','🍀','🌸','🌺','🌻','🌼',
            '🍁','🍂','🌾','🎋','🎍','🌵','🌊','🔥','⛅','🌈',
            # 음식
            '🍎','🍊','🍋','🍇','🍓','🫐','🍔','🍕','🍣','☕',
            # 업무·도구
            '📋','📝','📊','📅','📌','📍','🔧','⚙️','🔑','💡',
            '🎯','🚀','💪','🙌','👥','⭐','🏆','🎉','✅','⏰',
            '📞','📢','🔔','🏗️','💰','📦','🎨','🛠️','🔍','💼',
            '📁','🗂️','📮','📬','🖥️','⌨️','🖱️','📱','📷','🎥',
            # 스포츠·활동
            '⚽','🏀','🎾','🏊','🚴','🏋️','🎮','♟️','🎯','🏅',
            # 기호·색상
            '🔴','🟡','🟢','🔵','🟠','🟣','⚫','⚪','🔶','🔷',
            '⭐','💫','✨','❤️','🧡','💛','💚','💙','💜','🖤',
        ]
        ctx.update({
            'year': year, 'month': month,
            'prev_year': prev_year, 'prev_month': prev_month,
            'next_year': next_year, 'next_month': next_month,
            'cal': cal, 'today': today,
            'day_dots_json': day_dots_json,
            'selected_date': self.request.GET.get('selected', ''),
            'mode': mode,
            'emoji_list': emoji_list,
        })
        return ctx


@login_required
def task_day_tasks(request, date_str):
    """AJAX: 특정 날짜의 투두 목록 반환 (mode=team|personal)"""
    if _require_operations(request):
        return JsonResponse({'error': 'forbidden'}, status=403)
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        return JsonResponse({'error': 'invalid date'}, status=400)

    mode = request.GET.get('mode', 'team')

    if mode == 'personal':
        # 개인: start_date <= target <= end_date 범위에 걸치는 업무 모두 표시
        tasks_qs = (
            DailyTask.objects
            .filter(
                user=request.user,
                user__organization='operations',
                start_date__lte=target_date,
            )
            .filter(
                Q(end_date__gte=target_date) |
                Q(end_date__isnull=True, start_date=target_date)
            )
            .select_related('user')
            .prefetch_related('subtasks')
            .order_by('status', 'end_date', '-created_at')
        )
    else:
        tasks_qs = (
            DailyTask.objects
            .filter(start_date=target_date, user__organization='operations')
            .select_related('user')
            .prefetch_related('subtasks')
            .order_by('user__last_name', 'user__username', 'status', '-created_at')
        )

    my_tasks, team_tasks = [], []
    for t in tasks_qs:
        subtasks = [
            {'pk': st.pk, 'title': st.title, 'is_done': st.is_done}
            for st in t.subtasks.all()
        ]
        item = {
            'pk': t.pk,
            'task_name': t.task_name,
            'start_date': t.start_date.isoformat(),
            'end_date': t.end_date.isoformat() if t.end_date else '',
            'completed_date': t.completed_date.isoformat() if t.completed_date else '',
            'progress': t.progress,
            'status': t.status,
            'status_display': t.get_status_display(),
            'note': t.note,
            'user_name': t.user.get_full_name() or t.user.username,
            'user_emoji': t.user.emoji,
            'subtasks': subtasks,
        }
        (my_tasks if t.user_id == request.user.pk else team_tasks).append(item)

    return JsonResponse({'my_tasks': my_tasks, 'team_tasks': team_tasks})


class TaskManagerReportView(ManagerRequiredMixin, TemplateView):
    """관리자용: 날짜별 직원 업무 현황"""
    template_name = 'reports/task_manager_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_str = self.request.GET.get('date')
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            target_date = timezone.localdate()

        done_tasks = (
            DailyTask.objects
            .filter(completed_date=target_date, status='done')
            .select_related('user')
            .order_by('user__last_name', 'user__username')
        )
        pending_tasks = (
            DailyTask.objects
            .filter(start_date__lte=target_date)
            .exclude(status='done')
            .select_related('user')
            .order_by('user__last_name', 'user__username')
        )

        ctx['target_date'] = target_date
        ctx['prev_date'] = target_date - datetime.timedelta(days=1)
        ctx['next_date'] = target_date + datetime.timedelta(days=1)
        ctx['done_tasks'] = done_tasks
        ctx['pending_tasks'] = pending_tasks
        return ctx


@login_required
def task_daily_pdf(request):
    """일일 업무 일지 PDF 다운로드 (관리자용)"""
    if _require_operations(request) or request.user.role != 'manager':
        return redirect('main_menu')
    date_str = request.GET.get('date')
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except (TypeError, ValueError):
        target_date = timezone.localdate()

    tasks = (
        DailyTask.objects
        .filter(Q(completed_date=target_date) | Q(start_date__lte=target_date, status__in=('doing', 'hold')))
        .select_related('user')
        .order_by('user__last_name', 'user__username')
    )
    User = get_user_model()
    all_users = (
        User.objects
        .filter(is_active=True, role__in=('staff', 'up_staff'), organization='operations')
        .order_by('last_name', 'username')
    )
    pdf_bytes = build_daily_task_pdf(target_date, tasks, all_users=all_users)
    filename = f'daily_task_{target_date}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class TaskWeeklyReportView(ManagerRequiredMixin, TemplateView):
    """관리자용: 주간 투두 업무 현황"""
    template_name = 'reports/task_weekly_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        week_str = self.request.GET.get('week')
        try:
            year, week_num = week_str.split('-W')
            week_start = datetime.datetime.strptime(
                f'{year}-W{int(week_num)}-1', '%Y-W%W-%w').date()
        except (TypeError, ValueError, AttributeError):
            today = timezone.localdate()
            week_start = today - datetime.timedelta(days=today.weekday())

        week_end = week_start + datetime.timedelta(days=4)  # 월~금
        days = {}
        for d in range(5):
            day = week_start + datetime.timedelta(days=d)
            done = (
                DailyTask.objects
                .filter(completed_date=day, status='done')
                .select_related('user')
                .order_by('user__last_name', 'user__username')
            )
            pending = (
                DailyTask.objects
                .filter(start_date__lte=day)
                .exclude(status='done')
                .select_related('user')
                .order_by('user__last_name', 'user__username')
            )
            days[day] = {
                'done': done,
                'pending': pending,
                'total': done.count() + pending.count(),
            }

        prev_week = (week_start - datetime.timedelta(days=7)).strftime('%Y-W%W')
        next_week = (week_start + datetime.timedelta(days=7)).strftime('%Y-W%W')

        ctx['week_start'] = week_start
        ctx['week_end'] = week_end
        ctx['days'] = days
        ctx['prev_week'] = prev_week
        ctx['next_week'] = next_week
        ctx['current_week'] = week_start.strftime('%Y-W%W')
        return ctx


@login_required
def task_weekly_pdf(request):
    """주간 업무 일지 PDF 다운로드 (관리자용)"""
    if _require_operations(request) or request.user.role != 'manager':
        return redirect('main_menu')
    week_str = request.GET.get('week')
    try:
        year, week_num = week_str.split('-W')
        week_start = datetime.datetime.strptime(
            f'{year}-W{int(week_num)}-1', '%Y-W%W-%w').date()
    except (TypeError, ValueError, AttributeError):
        today = timezone.localdate()
        week_start = today - datetime.timedelta(days=today.weekday())

    week_end = week_start + datetime.timedelta(days=4)
    days = {}
    for d in range(5):
        day = week_start + datetime.timedelta(days=d)
        days[day] = (
            DailyTask.objects
            .filter(Q(completed_date=day) | Q(start_date__lte=day, status__in=('doing', 'hold')))
            .select_related('user')
            .order_by('user__last_name', 'user__username')
        )

    User = get_user_model()
    all_users = (
        User.objects
        .filter(is_active=True, role__in=('staff', 'up_staff'), organization='operations')
        .order_by('last_name', 'username')
    )
    pdf_bytes = build_weekly_task_pdf(week_start, week_end, days, all_users=all_users)
    filename = f'weekly_task_{week_start}_{week_end}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── 용산어린이정원 일일보고 ────────────────────────────────────────

class IntegratedDailyReportView(OperationsAccessMixin, TemplateView):
    """날짜별 운영데이터 입력 + 미리보기 페이지"""
    template_name = 'reports/integrated_daily.html'

    def _get_date(self):
        date_str = self.request.GET.get('date') or self.request.POST.get('date')
        try:
            return datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            return timezone.localdate()

    def get_context_data(self, **kwargs):
        from .weather import fetch_tomorrow_weather
        ctx = super().get_context_data(**kwargs)
        target_date = self._get_date()
        ops = OperationsDailyData.objects.filter(report_date=target_date).first()
        prev_date = target_date - datetime.timedelta(days=1)
        prev_ops = OperationsDailyData.objects.filter(report_date=prev_date).first()
        integrated = _gather_integrated_data(target_date)
        ctx['target_date']       = target_date
        ctx['prev_date']         = prev_date
        ctx['next_date']         = target_date + datetime.timedelta(days=1)
        ctx['ops']               = ops
        ctx['yesterday_total']   = prev_ops.today_total if prev_ops else 0
        ctx['weather_auto']      = fetch_tomorrow_weather(target_date)
        ctx['eoulrim_report']    = integrated['eoulrim_report']
        ctx['jamjam_report']     = integrated['jamjam_report']
        ctx['kumnare_report']    = integrated['kumnare_report']
        ctx['info_report']       = integrated['info_report']
        ctx['total_sales']       = integrated['total_sales']
        # 행사 + 작업사진 (카테고리별 분리)
        if ops:
            ctx['internal_events'] = [
                {'name': ev.name, 'columns': ev.columns_json or []}
                for ev in ops.internal_events.all().order_by('order')
            ]
            ctx['external_events'] = [
                {'name': ev.name, 'columns': ev.columns_json or []}
                for ev in ops.external_events.all().order_by('order')
            ]
            all_photos = list(ops.work_photos.all().order_by('order'))
            ctx['work_photos_interior'] = [p for p in all_photos if p.category == 'interior']
            ctx['work_photos_outdoor']  = [p for p in all_photos if p.category == 'outdoor']
            ctx['work_photos_fountain'] = [p for p in all_photos if p.category == 'fountain']
        else:
            ctx['internal_events']      = []
            ctx['external_events']      = []
            ctx['work_photos_interior'] = []
            ctx['work_photos_outdoor']  = []
            ctx['work_photos_fountain'] = []
        return ctx

    def post(self, request):
        target_date = self._get_date()
        ops, _ = OperationsDailyData.objects.get_or_create(
            report_date=target_date,
            defaults={'created_by': request.user},
        )
        # GODATA 자동수집으로 먼저 생성된 경우 created_by가 비어있음 → 첫 저장 시 채움
        if not ops.created_by_id:
            ops.created_by = request.user

        def _int(key, default=0):
            raw = request.POST.get(key, default)
            if isinstance(raw, str):
                raw = raw.replace(',', '').strip()
            try:
                return int(raw)
            except (ValueError, TypeError):
                return default

        ops.main_gate_walk   = _int('main_gate_walk')
        ops.sub_gate_walk    = _int('sub_gate_walk')
        ops.car_visit        = _int('car_visit')
        # 입장 총수 = 주출입구 + 부출입구 + 차량방문 (수기 수정값 반영)
        ops.today_total      = ops.main_gate_walk + ops.sub_gate_walk + ops.car_visit
        # 전일 입장 총수: 전날 today_total 자동 참조
        prev_ops = OperationsDailyData.objects.filter(
            report_date=target_date - datetime.timedelta(days=1)).first()
        ops.yesterday_total = prev_ops.today_total if prev_ops else 0
        ops.facility_interior = request.POST.get('facility_interior', '').strip()
        ops.facility_outdoor  = request.POST.get('facility_outdoor', '').strip()
        ops.facility_fountain = request.POST.get('facility_fountain', '').strip()
        ops.parking_family    = _int('parking_family')
        ops.parking_disabled  = _int('parking_disabled')
        ops.parking_pregnant  = _int('parking_pregnant')
        ops.parking_children  = _int('parking_children')
        ops.internal_event    = request.POST.get('internal_event', '').strip()
        ops.external_event    = request.POST.get('external_event', '').strip()
        ops.special_notes     = request.POST.get('special_notes', '').strip()
        ops.facility_interior_caption = request.POST.get('facility_interior_caption', '').strip()
        ops.facility_outdoor_caption  = request.POST.get('facility_outdoor_caption',  '').strip()
        ops.facility_fountain_caption = request.POST.get('facility_fountain_caption', '').strip()
        # 편익시설 매출 수기 입력 (해당 시설 보고서 없는 경우만 의미 있음)
        ops.manual_eoulrim_sales = _int('manual_eoulrim_sales')
        ops.manual_jamjam_sales  = _int('manual_jamjam_sales')
        ops.manual_kumnare_sales = _int('manual_kumnare_sales')
        # 세부 이용현황 수기 입력 (info/kumnare 보고서 없을 때)
        ops.manual_shuttle_total = _int('manual_shuttle_total')
        ops.manual_rental_total  = _int('manual_rental_total')
        ops.manual_stamp_total   = _int('manual_stamp_total')
        ops.save()

        # ── 내부/외부 행사: JSON 페이로드 받아서 전체 교체 ────────
        import json
        from .models import InternalEvent, ExternalEvent, FacilityWorkPhoto

        def _replace_events(model, related_mgr, json_key):
            raw = request.POST.get(json_key, '[]')
            try:
                items = json.loads(raw)
            except (ValueError, TypeError):
                items = []
            related_mgr.all().delete()
            for idx, ev in enumerate(items):
                name = (ev.get('name') or '').strip()
                cols = ev.get('columns') or []
                if not name and not cols:
                    continue
                model.objects.create(
                    ops=ops, name=name, columns_json=cols, order=idx,
                )

        _replace_events(InternalEvent, ops.internal_events, 'internal_events_json')
        _replace_events(ExternalEvent, ops.external_events, 'external_events_json')

        # ── 작업사진: 카테고리별 삭제 + 새로 업로드 ──────────────
        delete_ids = request.POST.getlist('delete_work_photo_ids')
        if delete_ids:
            ops.work_photos.filter(id__in=delete_ids).delete()

        for cat in ('interior', 'outdoor', 'fountain'):
            files = request.FILES.getlist(f'work_photos_{cat}')
            if not files:
                continue
            base_order = ops.work_photos.filter(category=cat).count()
            for i, f in enumerate(files):
                FacilityWorkPhoto.objects.create(
                    ops=ops, category=cat, image=f, order=base_order + i,
                )

        return redirect(f'/reports/integrated/?date={target_date}')


@login_required
def godata_manual_fetch(request):
    """GODATA 방문자 수 수동 수집 (UI 버튼에서 호출)."""
    if _require_operations(request):
        return redirect('main_menu')
    if request.method != 'POST':
        return redirect('reports:integrated_daily')

    date_str = request.POST.get('date')
    target_date = None
    if date_str:
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except ValueError:
            pass

    from .godata_scraper import sync_godata_to_db
    success = sync_godata_to_db(target_date)

    date_param = f'?date={target_date}' if target_date else ''
    if success:
        messages.success(request, 'GODATA 방문자 수 수집이 완료됐습니다.')
    else:
        messages.error(request, 'GODATA 수집 실패 — 서버 로그를 확인하세요.')

    return redirect(f'/reports/integrated/{date_param}')


@login_required
def integrated_daily_excel(request):
    """방문객 통계 엑셀 다운로드 — 기준 파일에 DB 신규 데이터 추가"""
    if _require_operations(request):
        return redirect('main_menu')

    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from pathlib import Path

    BASE_FILE = Path(settings.BASE_DIR) / 'reports' / 'excel_base' / '방문객통계_기준.xlsx'
    wb = openpyxl.load_workbook(BASE_FILE)
    ws = wb.active

    # ── 기준 파일 스캔: 날짜 + 마지막 데이터 행 + 합계 행(SUM 수식) ─
    existing_dates = set()
    last_data_row = 1   # 헤더(1행) 기준 시작
    sum_row = None      # 합계 행 (A열 비어있고 B열이 =SUM(...))
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        a_val = row[0]
        b_val = row[1] if len(row) > 1 else None
        if a_val is None:
            if isinstance(b_val, str) and b_val.upper().startswith('=SUM('):
                sum_row = r
                break  # 합계 행 이후는 무시
            continue
        last_data_row = r
        if hasattr(a_val, 'date'):     # datetime → date
            existing_dates.add(a_val.date())
        elif isinstance(a_val, datetime.date):
            existing_dates.add(a_val)

    # ── 기준 파일에 없는 날짜만 DB에서 조회 ──────────────────
    # 시간대별 데이터가 하나도 없는 행(slot 필드 전부 None)은 제외
    from django.db.models import Q
    slot_filter = Q()
    for h in range(9, 21):
        slot_filter |= Q(**{f'slot_{h:02d}00_main__isnull': False})
        slot_filter |= Q(**{f'slot_{h:02d}00_sub__isnull':  False})
    qs = OperationsDailyData.objects.exclude(
        report_date__in=existing_dates
    ).filter(slot_filter).order_by('report_date')

    if not qs.exists():
        # 추가할 데이터 없음 — 기준 파일 그대로 반환
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        today = timezone.localdate()
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        from urllib.parse import quote
        filename = f'(중앙일보)용산어린이정원_방문객 통계_{today:%y%m%d}.xlsx'
        response['Content-Disposition'] = (
            f"attachment; filename=\"visitor_stats_{today:%y%m%d}.xlsx\"; "
            f"filename*=UTF-8''{quote(filename, safe='')}"
        )
        return response

    # ── 데이터 행 스타일 참조 (2번째 행 기준) ─────────────────
    center = Alignment(horizontal='center', vertical='center')

    def _copy_style(src_cell, dst_cell):
        """폰트·정렬·테두리만 복사 (fill 제외 — 데이터행은 채우기 없음)"""
        if src_cell.font:
            dst_cell.font = src_cell.font.copy()
        if src_cell.border:
            dst_cell.border = src_cell.border.copy()
        dst_cell.alignment = center

    # 스타일 참조용 2행 셀 목록
    ref_row = 2

    def _slot(main, sub):
        if main is None and sub is None:
            return None
        return (main or 0) + (sub or 0)

    # ── 새 행 삽입 위치 결정 ──────────────────────────────────
    new_rows = list(qs)
    n = len(new_rows)

    if sum_row is not None:
        # 합계 행 위에 n개의 빈 행 삽입 → 합계 행은 sum_row+n으로 밀림
        ws.insert_rows(sum_row, amount=n)
        insert_start = sum_row
        new_sum_row  = sum_row + n
    else:
        insert_start = last_data_row + 1
        new_sum_row  = None

    # ── 새 행 채우기 ──────────────────────────────────────────
    for i, ops in enumerate(new_rows):
        r = insert_start + i

        row_values = [
            ops.report_date,
            f'=SUM(C{r}:M{r})',
            _slot(ops.slot_0900_main, ops.slot_0900_sub),
            _slot(ops.slot_1000_main, ops.slot_1000_sub),
            _slot(ops.slot_1100_main, ops.slot_1100_sub),
            _slot(ops.slot_1200_main, ops.slot_1200_sub),
            _slot(ops.slot_1300_main, ops.slot_1300_sub),
            _slot(ops.slot_1400_main, ops.slot_1400_sub),
            _slot(ops.slot_1500_main, ops.slot_1500_sub),
            _slot(ops.slot_1600_main, ops.slot_1600_sub),
            _slot(ops.slot_1700_main, ops.slot_1700_sub),
            _slot(ops.slot_1800_main, ops.slot_1800_sub),
            _slot(ops.slot_1900_main, ops.slot_1900_sub),
            ops.car_visit or '-',
            ops.main_gate_walk,
            ops.sub_gate_walk,
        ]

        for col_idx, val in enumerate(row_values, start=1):
            dst = ws.cell(row=r, column=col_idx, value=val)
            src = ws.cell(row=ref_row, column=col_idx)
            _copy_style(src, dst)
            if col_idx == 1:
                dst.number_format = 'mm-dd-yy'

    # ── 합계 행 SUM 수식 범위 갱신 (B~P열) ────────────────────
    if new_sum_row is not None:
        last_data_after = new_sum_row - 1
        for c in range(2, 17):  # B(2) ~ P(16)
            col_letter = openpyxl.utils.get_column_letter(c)
            ws.cell(row=new_sum_row, column=c).value = (
                f'=SUM({col_letter}2:{col_letter}{last_data_after})'
            )

    # ── 반환 ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = timezone.localdate()
    from urllib.parse import quote
    filename = f'(중앙일보)용산어린이정원_방문객 통계_{today:%y%m%d}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f"attachment; filename=\"visitor_stats_{today:%y%m%d}.xlsx\"; "
        f"filename*=UTF-8''{quote(filename, safe='')}"
    )
    return response


def _sf_slot(sf_reservations, sf_entries, field_types, start_time):
    """구장/시작시간으로 entry 또는 reservation 데이터를 반환한다."""
    CAT = {'normal': '일반', 'quarter': '쿼터', 'event': '행사', 'other': '기타'}
    if isinstance(field_types, str):
        field_types = [field_types]
    for e in sf_entries:
        if e.field_type in field_types and e.time_start == start_time:
            # 예약인원: entry의 reserved_* 필드 우선, 없으면 매칭 Reservation의 total_users
            res = None
            if e.reserved_adult_count is not None or e.reserved_child_count is not None:
                res = (e.reserved_adult_count or 0) + (e.reserved_child_count or 0)
            else:
                rv = next((r for r in sf_reservations
                           if r.field_type in field_types and r.time_start == start_time), None)
                res = rv.total_users if rv and rv.total_users else None
            # 입장인원: actual_* 필드
            act = None
            if e.actual_adult_count is not None:
                act = (e.actual_adult_count or 0) + (e.actual_child_count or 0)
            return {
                'cat':      CAT.get(e.category, e.category),
                'reserved': res,
                'actual':   act,
            }
    for r in sf_reservations:
        if r.field_type in field_types and r.time_start == start_time:
            return {
                'cat':      '일반',
                'reserved': r.total_users,
                'actual':   r.actual_adult_count,
            }
    return {'cat': None, 'reserved': None, 'actual': None}


def _sf_day_total_by_cat(sf_reservations, sf_entries, cat):
    """하루 전체 모든 구장의 분류별(일반/쿼터) 합계를 반환한다."""
    total_res = 0
    total_act = 0

    # 쿼터 entry 키 (field_type, time_start) 집합
    quarter_keys = {(e.field_type, e.time_start) for e in sf_entries if e.category == 'quarter'}
    # entry가 있는 키 집합
    entry_keys = {(e.field_type, e.time_start) for e in sf_entries}

    if cat == '쿼터':
        for e in sf_entries:
            if e.category == 'quarter':
                # 예약인원: entry의 reserved_* 우선, 없으면 매칭 Reservation
                if e.reserved_adult_count is not None or e.reserved_child_count is not None:
                    total_res += (e.reserved_adult_count or 0) + (e.reserved_child_count or 0)
                else:
                    rv = next((r for r in sf_reservations
                               if r.field_type == e.field_type and r.time_start == e.time_start), None)
                    total_res += rv.total_users or 0 if rv else 0
                # 입장인원
                total_act += (e.actual_adult_count or 0) + (e.actual_child_count or 0)
    else:  # 일반
        for e in sf_entries:
            if e.category != 'quarter':
                if e.reserved_adult_count is not None or e.reserved_child_count is not None:
                    total_res += (e.reserved_adult_count or 0) + (e.reserved_child_count or 0)
                total_act += (e.actual_adult_count or 0) + (e.actual_child_count or 0)
        for r in sf_reservations:
            if (r.field_type, r.time_start) not in quarter_keys:
                if (r.field_type, r.time_start) not in entry_keys:
                    # entry 없는 일반 예약: reservation 기준
                    total_res += r.total_users or 0
                    total_act += r.actual_adult_count or 0

    return {
        'cat':      cat,
        'reserved': total_res or None,
        'actual':   total_act or None,
    }


# 합계 행 분류 고정: 1타임=일반, 2타임=쿼터
_BB_TOTAL_CATS = ['일반', '쿼터']


def _organize_sf_slots(sf_reservations, sf_entries):
    """스포츠필드 예약/이용 데이터를 구장·타임별 구조로 정리한다."""
    import datetime as dt

    # 축구장·테니스장 공통 3타임
    ST = [
        {'label': '1타임 (10:00~12:00)', 'start': dt.time(10, 0),  'end': dt.time(12, 0)},
        {'label': '2타임 (13:00~15:00)', 'start': dt.time(13, 0),  'end': dt.time(15, 0)},
        {'label': '3타임 (15:30~17:30)', 'start': dt.time(15, 30), 'end': dt.time(17, 30)},
    ]
    # 야구장 2타임
    BB = [
        {'label': '1타임 (10:00~14:00)', 'start': dt.time(10, 0), 'end': dt.time(14, 0)},
        {'label': '2타임 (14:00~18:00)', 'start': dt.time(14, 0), 'end': dt.time(18, 0)},
    ]

    st_rows = [
        {
            'label':  s['label'],
            'soccer': _sf_slot(sf_reservations, sf_entries, 'soccer', s['start']),
            'tennis': _sf_slot(sf_reservations, sf_entries,
                               ['tennis_grass', 'tennis_hard'], s['start']),
        }
        for s in ST
    ]
    # 합계는 하루 전체 기준으로 분류별 집계 (1타임=일반, 2타임=쿼터)
    day_totals = {cat: _sf_day_total_by_cat(sf_reservations, sf_entries, cat)
                  for cat in _BB_TOTAL_CATS}

    bb_rows = [
        {
            'label':    s['label'],
            'baseball': _sf_slot(sf_reservations, sf_entries, 'baseball', s['start']),
            'total':    day_totals[_BB_TOTAL_CATS[i]],
        }
        for i, s in enumerate(BB)
    ]
    return {'st_rows': st_rows, 'bb_rows': bb_rows}


def _gather_integrated_data(target_date):
    """통합 일일보고에 필요한 모든 데이터를 수집한다."""
    from sportsfield.models import Reservation, SportsfieldEntry
    from facilities.models import EoulrimReport, JamjamReport, KumnareReport
    from info.models import InfoReport

    ops_data        = OperationsDailyData.objects.filter(report_date=target_date).first()
    sf_reservations = list(Reservation.objects.filter(
        reservation_date=target_date, status='confirmed').order_by('field_type', 'time_start'))
    sf_entries      = list(SportsfieldEntry.objects.filter(
        entry_date=target_date).order_by('field_type', 'time_start'))
    eoulrim_report  = EoulrimReport.objects.filter(report_date=target_date).first()
    jamjam_report   = JamjamReport.objects.filter(report_date=target_date).first()
    kumnare_report  = KumnareReport.objects.filter(report_date=target_date).first()
    info_report     = InfoReport.objects.prefetch_related('items').filter(
        report_date=target_date).first()

    info_shuttle_items = list(info_report.items.filter(section='shuttle').order_by('order')) if info_report else []
    info_patrol_items  = list(info_report.items.filter(section='patrol').order_by('order'))  if info_report else []

    # 시설 보고서가 없으면 OperationsDailyData의 수기 매출 사용
    e_s = (eoulrim_report.daily_net_sales if eoulrim_report
           else (ops_data.manual_eoulrim_sales if ops_data else 0))
    j_s = (jamjam_report.daily_net_sales  if jamjam_report
           else (ops_data.manual_jamjam_sales  if ops_data else 0))
    k_s = (kumnare_report.sales_amount    if kumnare_report
           else (ops_data.manual_kumnare_sales if ops_data else 0))

    return {
        'ops_data':            ops_data,
        'sf_reservations':     sf_reservations,
        'sf_entries':          sf_entries,
        'sf_slots':            _organize_sf_slots(sf_reservations, sf_entries),
        'eoulrim_report':      eoulrim_report,
        'jamjam_report':       jamjam_report,
        'kumnare_report':      kumnare_report,
        'info_report':         info_report,
        'info_shuttle_items':  info_shuttle_items,
        'info_patrol_items':   info_patrol_items,
        'total_sales':         e_s + j_s + k_s,
        'eoulrim_sales':       e_s,
        'jamjam_sales':        j_s,
        'kumnare_sales':       k_s,
    }


@login_required
def integrated_daily_pdf(request):
    """용산어린이정원 일일보고 PDF 다운로드 (WeasyPrint)"""
    if _require_operations(request):
        return redirect('main_menu')

    date_str = request.GET.get('date')
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except (TypeError, ValueError):
        target_date = timezone.localdate()
    data = _gather_integrated_data(target_date)
    od   = data['ops_data']

    from django.template.loader import render_to_string
    from weasyprint import HTML as WeasyHTML

    ctx = {
        'target_date':        target_date,
        'ops':                od,
        'sf_slots':           data['sf_slots'],
        'eoulrim':            data['eoulrim_report'],
        'jamjam':             data['jamjam_report'],
        'kumnare':            data['kumnare_report'],
        'info':               data['info_report'],
        'info_shuttle_items': data['info_shuttle_items'],
        'info_patrol_items':  data['info_patrol_items'],
        'total_sales':        data['total_sales'],
    }
    html_string = render_to_string('reports/integrated_daily_pdf.html', ctx)
    pdf_bytes   = WeasyHTML(string=html_string).write_pdf()

    filename = f'용산어린이정원_일일보고_{target_date}.pdf'
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _validate_for_download(ops, eoulrim, jamjam, kumnare, info):
    """한글 다운로드 전 필수 항목 입력 검증. 누락 메시지 리스트 반환."""
    errors = []
    if not ops or not ops.pk:
        errors.append('일일보고 데이터가 없습니다. 입력 후 저장해주세요.')
        return errors
    if not ops.created_by_id:
        errors.append('저장 버튼을 먼저 눌러주세요. (작성자 정보 없음)')
        return errors

    # 1. 금일 방문현황
    if (ops.main_gate_walk + ops.sub_gate_walk + ops.car_visit) == 0:
        errors.append('"금일 방문현황"이 입력되지 않았습니다.')
    # 2. 주차장 (대수)
    if (ops.parking_family + ops.parking_disabled + ops.parking_pregnant + ops.parking_children) == 0:
        errors.append('"주차장(대수)"이 입력되지 않았습니다.')
    # 3. 편익시설 매출 (자동 OR 수기 합계)
    e = (eoulrim.daily_net_sales if eoulrim else 0) or ops.manual_eoulrim_sales
    j = (jamjam.daily_net_sales  if jamjam  else 0) or ops.manual_jamjam_sales
    k = (kumnare.sales_amount    if kumnare else 0) or ops.manual_kumnare_sales
    if (e + j + k) == 0:
        errors.append('"편익시설 매출"이 입력되지 않았습니다.')
    # 4. 세부 이용현황
    s  = (info.shuttle_total if info else 0) or ops.manual_shuttle_total
    r  = (kumnare.rental_total_users if kumnare else 0) or ops.manual_rental_total
    st = (kumnare.stamp_issued       if kumnare else 0) or ops.manual_stamp_total
    if (s + r + st) == 0:
        errors.append('"세부 이용현황"이 입력되지 않았습니다.')

    return errors


@login_required
def integrated_daily_hwp(request):
    """용산어린이정원 일일보고 한글파일 다운로드"""
    if _require_operations(request):
        return redirect('main_menu')

    date_str = request.GET.get('date')
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except (TypeError, ValueError):
        target_date = timezone.localdate()

    data = _gather_integrated_data(target_date)

    # ── 필수 항목 검증 (값 없거나 미저장 시 차단) ───────────
    errors = _validate_for_download(
        data['ops_data'], data['eoulrim_report'],
        data['jamjam_report'], data['kumnare_report'], data['info_report'],
    )
    if errors:
        from django.contrib import messages
        for e in errors:
            messages.error(request, e)
        return redirect(f'/reports/integrated/?date={target_date}')

    from .weather import fetch_tomorrow_weather
    ops = data['ops_data']
    if ops is None:
        ops = OperationsDailyData(report_date=target_date)

    # 전일 입장 총수 자동 참조
    prev_ops = OperationsDailyData.objects.filter(
        report_date=target_date - datetime.timedelta(days=1)).first()
    ops.yesterday_total = prev_ops.today_total if prev_ops else 0

    # 명일 기상 항상 API에서 자동 조회
    weather = fetch_tomorrow_weather(target_date)
    if weather:
        ops.tomorrow_temp_min = weather['temp_min']
        ops.tomorrow_temp_max = weather['temp_max']
        ops.tomorrow_rain_pct = weather['rain_pct']

    from .hwpx_report import build_integrated_daily_hwpx
    hwpx_bytes = build_integrated_daily_hwpx(
        target_date        = target_date,
        ops                = ops,
        sf_slots           = data['sf_slots'],
        eoulrim            = data['eoulrim_report'],
        jamjam             = data['jamjam_report'],
        kumnare            = data['kumnare_report'],
        info_report        = data['info_report'],
        info_shuttle_items = data['info_shuttle_items'],
        info_patrol_items  = data['info_patrol_items'],
        total_sales        = data['total_sales'],
    )
    from urllib.parse import quote
    """파일 명칭 변환"""
    target_date_set = f"{target_date:%y%m%d}"
    filename = f'(중앙일보)용산어린이정원_일일보고_{target_date_set}.hwpx'
    encoded  = quote(filename, safe='')
    response = HttpResponse(hwpx_bytes, content_type='application/hwp+zip')
    response['Content-Disposition'] = (
        f"attachment; filename=\"daily_report_{target_date_set}.hwpx\"; "
        f"filename*=UTF-8''{encoded}"
    )
    return response


class DailyDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'reports/dashboard_daily.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_str = self.request.GET.get('date')
        try:
            target_date = datetime.date.fromisoformat(date_str)
        except (TypeError, ValueError):
            target_date = timezone.localdate()

        reports = (
            DailyReport.objects
            .filter(report_date=target_date)
            .select_related('author')
            .prefetch_related('task_items')
        )
        ctx['target_date'] = target_date
        ctx['prev_date'] = target_date - datetime.timedelta(days=1)
        ctx['next_date'] = target_date + datetime.timedelta(days=1)
        ctx['reports'] = reports
        return ctx


class WeeklyDashboardView(ManagerRequiredMixin, TemplateView):
    template_name = 'reports/dashboard_weekly.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        week_str = self.request.GET.get('week')
        try:
            # week 형식: "2026-W14"
            year, week_num = week_str.split('-W')
            week_start = datetime.datetime.strptime(f'{year}-W{int(week_num)}-1', '%Y-W%W-%w').date()
        except (TypeError, ValueError, AttributeError):
            today = timezone.localdate()
            week_start = today - datetime.timedelta(days=today.weekday())

        week_end = week_start + datetime.timedelta(days=6)
        prev_week_start = week_start - datetime.timedelta(days=7)
        next_week_start = week_start + datetime.timedelta(days=7)

        reports = (
            DailyReport.objects
            .filter(report_date__range=(week_start, week_end))
            .select_related('author')
            .prefetch_related('task_items')
            .order_by('report_date', 'author__last_name')
        )

        # 날짜별로 그룹핑
        days = {}
        for d in range(7):
            day = week_start + datetime.timedelta(days=d)
            days[day] = []
        for report in reports:
            if report.report_date in days:
                days[report.report_date].append(report)

        ctx['week_start'] = week_start
        ctx['week_end'] = week_end
        ctx['days'] = days
        ctx['prev_week'] = prev_week_start.strftime('%Y-W%W')
        ctx['next_week'] = next_week_start.strftime('%Y-W%W')
        ctx['current_week'] = week_start.strftime('%Y-W%W')
        return ctx
